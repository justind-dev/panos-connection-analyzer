"""
reporter.py
-----------
Formats and writes connection analysis results to Excel or CSV.

Usage:
    reporter = Reporter()
    reporter.generate(connections, "report.xlsx")
"""

from typing import List

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class Reporter:
    def generate(self, connections: List[dict], output_file: str):
        if not connections:
            print("No connections to report.")
            return

        df = pd.DataFrame(connections)

        # Sanitize hostnames
        for col in ['source_hostname', 'destination_hostname']:
            if col in df.columns:
                df[col] = df[col].fillna('NotFound').astype(str).replace(['nan', '', 'None'], 'NotFound')

        # Format timestamps
        for col in ['first_seen', 'last_seen']:
            if col in df.columns:
                df[col] = df[col].astype(str)

        # Column order - include new NAT and address object fields
        preferred = [
            'direction', 'from_zone', 'to_zone',
            'source_ip', 'source_hostname', 'source_addr_objects', 'source_addr_groups',
            'destination_ip', 'destination_hostname', 'destination_addr_objects', 'destination_addr_groups',
            'nat_source_ip', 'nat_destination_ip',
            'protocol', 'port',
            'nat_source_port', 'nat_destination_port',
            'application',
            'app_category', 'app_technology', 'app_risk',
            'rule', 'nat_rule', 'nat_type',
            'action', 'hits',
            'bytes', 'bytes_sent', 'bytes_received', 'packets',
            'first_seen', 'last_seen'
        ]

        # Only include columns that actually exist in the data
        df = df[[col for col in preferred if col in df.columns]]

        # Sort
        sort_cols = ['direction', 'hits'] if 'direction' in df.columns else ['hits']
        df = df.sort_values(sort_cols, ascending=[True, False] if len(sort_cols) > 1 else [False])

        # Output
        if output_file.endswith(".csv"):
            df.to_csv(output_file, index=False)
        else:
            if not output_file.endswith(".xlsx"):
                output_file += ".xlsx"
            self._write_excel(df, output_file)

        print(f"\n✅ Report written to: {output_file}")
        print(f"🔹 {len(df)} unique connections\n")

        self._print_summary(df)

    def _write_excel(self, df: pd.DataFrame, path: str):
        writer = pd.ExcelWriter(path, engine='openpyxl')
        df.to_excel(writer, sheet_name="Connections", index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        ws = writer.sheets["Connections"]

        # Add table style
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)

        # Define table range
        table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        table = Table(displayName="ConnectionsTable", ref=table_range)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-width
        for idx, col in enumerate(df.columns, 1):
            width = min(max(len(col), df[col].astype(str).map(len).max()), 50)
            ws.column_dimensions[get_column_letter(idx)].width = width + 2

        # Header styling
        font = Font(bold=True)
        fill = PatternFill(start_color="D9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Add NAT summary sheet if NAT information exists
        self._add_nat_summary_sheet(workbook, df)

        # Add address objects summary sheet if address object information exists
        self._add_address_summary_sheet(workbook, df)

        writer.close()

    def _add_nat_summary_sheet(self, workbook, df: pd.DataFrame):
        """Add a NAT Summary sheet if NAT information exists"""

        # Check if we have any NAT information
        nat_columns = ['nat_rule', 'nat_type', 'nat_source_ip', 'nat_destination_ip']
        if not any(col in df.columns for col in nat_columns):
            return  # No NAT information, skip this sheet

        if 'nat_rule' in df.columns and not df['nat_rule'].notna().any():
            return  # NAT rules column exists but all values are null

        # Create a new sheet for NAT summary
        nat_sheet = workbook.create_sheet(title="NAT Summary")

        # Set up the headers
        nat_sheet['A1'] = "NAT Summary"
        nat_sheet['A1'].font = Font(bold=True, size=14)

        # Add NAT rule counts
        row = 3
        if 'nat_rule' in df.columns and df['nat_rule'].notna().any():
            nat_sheet[f'A{row}'] = "NAT Rules Used"
            nat_sheet[f'A{row}'].font = Font(bold=True)
            row += 1

            # Count connections per NAT rule
            nat_rules = df[df['nat_rule'].notna() & (df['nat_rule'] != '')].groupby('nat_rule').size().reset_index(
                name='count')
            nat_rules = nat_rules.sort_values('count', ascending=False)

            # Write rule counts
            for idx, (rule, count) in enumerate(zip(nat_rules['nat_rule'], nat_rules['count']), 1):
                nat_sheet[f'A{row}'] = rule
                nat_sheet[f'B{row}'] = count
                nat_sheet[f'C{row}'] = f"{count / len(df) * 100:.1f}% of connections"
                row += 1

            row += 2

        # Add NAT type summary
        if 'nat_type' in df.columns and df['nat_type'].notna().any():
            nat_sheet[f'A{row}'] = "NAT Types Used"
            nat_sheet[f'A{row}'].font = Font(bold=True)
            row += 1

            # Count connections per NAT type
            nat_types = df[df['nat_type'].notna() & (df['nat_type'] != '')].groupby('nat_type').size().reset_index(
                name='count')
            nat_types = nat_types.sort_values('count', ascending=False)

            # Write type counts
            for idx, (type_name, count) in enumerate(zip(nat_types['nat_type'], nat_types['count']), 1):
                nat_sheet[f'A{row}'] = type_name
                nat_sheet[f'B{row}'] = count
                nat_sheet[f'C{row}'] = f"{count / len(df) * 100:.1f}% of connections"
                row += 1

        # Auto-size columns
        for col in ['A', 'B', 'C']:
            max_length = 0
            for cell in nat_sheet[col]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            nat_sheet.column_dimensions[col].width = max_length + 2

    def _add_address_summary_sheet(self, workbook, df: pd.DataFrame):
        """Add an Address Objects summary sheet if address information exists"""

        # Check if we have any address object information
        addr_columns = ['source_addr_objects', 'source_addr_groups',
                        'destination_addr_objects', 'destination_addr_groups']
        if not any(col in df.columns for col in addr_columns):
            return  # No address information, skip this sheet

        if all(not df[col].notna().any() for col in addr_columns if col in df.columns):
            return  # Address columns exist but all values are null

        # Create a new sheet for address summary
        addr_sheet = workbook.create_sheet(title="Address Objects")

        # Set up the headers
        addr_sheet['A1'] = "Address Objects and Groups Summary"
        addr_sheet['A1'].font = Font(bold=True, size=14)

        row = 3

        # Helper function to add a section for a specific address type
        def add_address_section(column, title):
            nonlocal row
            if column in df.columns and df[column].notna().any():
                addr_sheet[f'A{row}'] = title
                addr_sheet[f'A{row}'].font = Font(bold=True)
                row += 1

                # Split the semicolon-separated list into individual objects
                all_items = []
                for items_str in df[column].dropna():
                    if items_str:
                        all_items.extend([item.strip() for item in items_str.split(';')])

                # Count frequency of each item
                from collections import Counter
                counts = Counter(all_items)

                # Sort by frequency
                for item, count in sorted(counts.items(), key=lambda x: x[1], reverse=True):
                    addr_sheet[f'A{row}'] = item
                    addr_sheet[f'B{row}'] = count
                    addr_sheet[f'C{row}'] = f"{count / len(df) * 100:.1f}% of connections"
                    row += 1

                row += 2  # Add space before next section

        # Add sections for each type of address object
        add_address_section('source_addr_objects', "Source Address Objects")
        add_address_section('source_addr_groups', "Source Address Groups")
        add_address_section('destination_addr_objects', "Destination Address Objects")
        add_address_section('destination_addr_groups', "Destination Address Groups")

        # Auto-size columns
        for col in ['A', 'B', 'C']:
            max_length = 0
            for cell in addr_sheet[col]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            addr_sheet.column_dimensions[col].width = max_length + 2

    def _print_summary(self, df: pd.DataFrame):
        if 'application' in df.columns:
            print("Top Applications:")
            print(df.groupby("application")["hits"].sum().sort_values(ascending=False).head(10))

        if 'from_zone' in df.columns and 'to_zone' in df.columns:
            print("\nZone Summary:")
            print(df.groupby(["from_zone", "to_zone"])["hits"].sum().sort_values(ascending=False).head(10))

        if 'action' in df.columns:
            print("\nActions:")
            print(df.groupby("action")["hits"].sum().sort_values(ascending=False))

        # Add NAT summary if NAT fields exist
        nat_fields = ['nat_source_ip', 'nat_destination_ip', 'nat_rule']
        if any(field in df.columns for field in nat_fields):
            print("\nNAT Usage:")

            # Count connections with NAT
            nat_ip_count = df[
                (df[nat_fields[0]].notna() | df[nat_fields[1]].notna()) if nat_fields[0] in df.columns and nat_fields[
                    1] in df.columns else False].shape[0]
            if nat_ip_count > 0:
                print(f"Connections using NAT IPs: {nat_ip_count} ({nat_ip_count / len(df) * 100:.1f}% of total)")

            # Add NAT rule summary if available
            if 'nat_rule' in df.columns and df['nat_rule'].notna().any():
                nat_rule_count = df[df['nat_rule'].notna() & (df['nat_rule'] != '')].shape[0]
                print(
                    f"Connections with identified NAT rules: {nat_rule_count} ({nat_rule_count / len(df) * 100:.1f}% of total)")

                print("\nTop NAT Rules:")
                nat_rules = df[df['nat_rule'].notna() & (df['nat_rule'] != '')]
                if not nat_rules.empty:
                    print(nat_rules.groupby("nat_rule")["hits"].sum().sort_values(ascending=False).head(10))

            # Add NAT type summary if available
            if 'nat_type' in df.columns and df['nat_type'].notna().any():
                print("\nNAT Types:")
                nat_types = df[df['nat_type'].notna() & (df['nat_type'] != '')]
                if not nat_types.empty:
                    print(nat_types.groupby("nat_type")["hits"].sum().sort_values(ascending=False))

        # Add address object summary if available
        addr_fields = ['source_addr_objects', 'source_addr_groups',
                       'destination_addr_objects', 'destination_addr_groups']
        if any(field in df.columns for field in addr_fields):
            addr_count = df[df[[field for field in addr_fields if field in df.columns]].notna().any(axis=1)].shape[0]
            if addr_count > 0:
                print("\nAddress Object Usage:")
                print(
                    f"Connections with identified address objects: {addr_count} ({addr_count / len(df) * 100:.1f}% of total)")
