#!/usr/bin/env python3
"""
PAN-OS Connection Analyzer with Optimized Hostname Filtering
-----------------------------------------------------------
Analyzes Palo Alto Networks traffic logs and allows filtering by
IP address, zone, or hostname with efficient DNS resolution.
"""

import argparse
import datetime
import socket
import sys
import concurrent.futures
from collections import defaultdict

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Required packages not installed. Please install with:")
    print("pip install pandas openpyxl")
    sys.exit(1)

class PANOSConnectionAnalyzer:
    def __init__(self, target_ip=None, target_zone=None, target_hostname=None, dns_timeout=2.0, max_workers=50):
        """Initialize the connection analyzer with targets."""
        self.target_ip = target_ip
        self.target_zone = target_zone
        self.target_hostname = target_hostname
        self.dns_timeout = dns_timeout  # Timeout for DNS lookups in seconds
        self.max_workers = max_workers  # Max parallel workers for DNS resolution
        self.dns_cache = {}  # Cache for DNS lookups
        
        # If hostname is provided, resolve it to an IP
        if target_hostname:
            resolved_ip = self.resolve_hostname_to_ip(target_hostname)
            if resolved_ip:
                print(f"Resolved hostname '{target_hostname}' to IP: {resolved_ip}")
                self.target_ip = resolved_ip
                self.analysis_mode = "ip"
            else:
                print(f"WARNING: Could not resolve hostname '{target_hostname}' to an IP address.")
                print("Will try to look for partial hostname matches in the traffic logs.")
                self.analysis_mode = "hostname"
        elif target_ip:
            self.analysis_mode = "ip"
        elif target_zone:
            self.analysis_mode = "zone"
        else:
            self.analysis_mode = "all"
    
    def resolve_hostname_to_ip(self, hostname):
        """Resolve a hostname to an IP address."""
        old_timeout = socket.getdefaulttimeout()
        socket.setdefaulttimeout(self.dns_timeout)
        
        try:
            # Try to get the IP address for the hostname
            ip_address = socket.gethostbyname(hostname)
            return ip_address
        except socket.gaierror:
            # No IP found for this hostname
            return None
        finally:
            socket.setdefaulttimeout(old_timeout)
    
    def resolve_dns_with_timeout(self, ip):
        """Resolve an IP address to a hostname with timeout."""
        if not ip or ip == "any" or ip == "":
            return "N/A"
        
        # Check cache first
        if ip in self.dns_cache:
            return self.dns_cache[ip]
        
        # Set socket timeout for this lookup
        old_timeout = socket.getdefaulttimeout()
        socket.setdefaulttimeout(self.dns_timeout)
        
        try:
            hostname, _, _ = socket.gethostbyaddr(ip)
            self.dns_cache[ip] = hostname
            return hostname
        except:
            # Any error returns "NotFound"
            self.dns_cache[ip] = "NotFound"
            return "NotFound"
        finally:
            # Restore original timeout
            socket.setdefaulttimeout(old_timeout)

    def resolve_ips_parallel(self, unique_ips):
        """Resolve multiple IPs in parallel with timeouts."""
        if not unique_ips:
            return {}
            
        # Define worker function for each IP
        def resolve_worker(ip):
            return ip, self.resolve_dns_with_timeout(ip)
        
        results = {}
        
        # Use ThreadPoolExecutor for parallel processing
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks and collect futures
            future_to_ip = {executor.submit(resolve_worker, ip): ip for ip in unique_ips}
            
            # Process results as they complete (with progress indicator)
            completed = 0
            total = len(future_to_ip)
            print(f"Starting parallel DNS resolution for {total} IPs...")
            
            for future in concurrent.futures.as_completed(future_to_ip):
                ip, hostname = future.result()
                results[ip] = hostname
                
                # Show progress
                completed += 1
                if completed % 20 == 0 or completed == total:
                    print(f"Resolved {completed}/{total} IPs")
        
        return results
    
    def find_partial_hostname_matches(self, logs_df):
        """Find all IPs with hostnames containing the target hostname pattern."""
        matching_ips = set()
        
        # Get all unique source and destination IPs
        if 'source_ip' in logs_df.columns:
            source_ips = set(logs_df['source_ip'].dropna().unique())
        else:
            source_ips = set()
            
        if 'destination_ip' in logs_df.columns:
            dest_ips = set(logs_df['destination_ip'].dropna().unique())
        else:
            dest_ips = set()
            
        all_ips = source_ips.union(dest_ips)
        print(f"Searching {len(all_ips)} unique IPs for partial hostname matches to '{self.target_hostname}'")
        
        # Resolve all IPs to hostnames
        hostname_map = self.resolve_ips_parallel(all_ips)
        self.dns_cache.update(hostname_map)
        
        # Find IPs that match the target hostname
        for ip, hostname in hostname_map.items():
            if self.target_hostname.lower() in hostname.lower():
                matching_ips.add(ip)
                
        print(f"Found {len(matching_ips)} IPs with hostnames containing '{self.target_hostname}'")
        if matching_ips:
            print("Matching IPs:", ", ".join(list(matching_ips)[:10]), 
                  "..." if len(matching_ips) > 10 else "")
                
        return matching_ips
    
    def analyze_logs_from_csv(self, csv_file):
        """Read and filter traffic logs from a CSV file."""
        try:
            print(f"Reading logs from {csv_file}...")
            df = pd.read_csv(csv_file)
            
            # Debug: Print column names to help diagnose issues
            print(f"CSV columns found: {', '.join(df.columns)}")
            
            # Standardize column names - map common PAN-OS column names
            column_mapping = {
                'Source address': 'source_ip',
                'Destination address': 'destination_ip',
                'IP Protocol': 'protocol',
                'Destination Port': 'port',
                'Application': 'application',
                'Action': 'action',
                'Source Zone': 'from_zone',
                'Destination Zone': 'to_zone',
                'Receive Time': 'receive_time',
                'Bytes': 'bytes',
                'Rule': 'rule',
                'Bytes Sent': 'bytes_sent',
                'Bytes Received': 'bytes_received',
                'Packets': 'packets',
                'Category of app': 'app_category',
                'Technology of app': 'app_technology',
                'Risk of app': 'app_risk',
                'Source Hostname': 'source_hostname',
                'Destination Hostname': 'destination_hostname'
            }
            
            # Rename columns that exist in the DataFrame
            df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
            
            # Handle hostname-based filtering (partial matches)
            if self.analysis_mode == "hostname":
                matching_ips = self.find_partial_hostname_matches(df)
                if matching_ips:
                    filtered_df = df[(df['source_ip'].isin(matching_ips)) | (df['destination_ip'].isin(matching_ips))]
                    print(f"Found {len(filtered_df)} log entries for hostname {self.target_hostname}")
                else:
                    filtered_df = pd.DataFrame()  # Empty DataFrame if no matching IPs
                    print(f"No log entries found for hostname {self.target_hostname}")
            
            # For other analysis modes, filter directly
            elif self.analysis_mode == "ip" and self.target_ip:
                filtered_df = df[(df['source_ip'] == self.target_ip) | (df['destination_ip'] == self.target_ip)]
                print(f"Found {len(filtered_df)} log entries for IP {self.target_ip}")
            elif self.analysis_mode == "zone" and self.target_zone:
                filtered_df = df[(df['from_zone'] == self.target_zone) | (df['to_zone'] == self.target_zone)]
                print(f"Found {len(filtered_df)} log entries for zone {self.target_zone}")
            else:
                filtered_df = df
                print(f"Analyzing all {len(filtered_df)} log entries")
            
            # Remove rows with NaN in critical columns
            required_cols = ['source_ip', 'destination_ip', 'protocol', 'port']
            existing_cols = [col for col in required_cols if col in filtered_df.columns]
            if existing_cols:
                filtered_df = filtered_df.dropna(subset=existing_cols)
            
            return filtered_df
            
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            sys.exit(1)
    
    def analyze_connections(self, logs_df):
        """Analyze logs to identify unique connections with parallel DNS resolution."""
        # Define default connection template with explicit string "NotFound" for hostnames
        def connection_template():
            return {
                'source_ip': '',
                'destination_ip': '',
                'source_hostname': 'NotFound',  # Explicit string value, not empty or None
                'destination_hostname': 'NotFound',  # Explicit string value, not empty or None
                'protocol': '',
                'port': '',
                'application': '',
                'app_category': '',
                'app_technology': '',
                'app_risk': '',
                'rule': '',
                'hits': 0,
                'bytes': 0,
                'bytes_sent': 0,
                'bytes_received': 0,
                'packets': 0,
                'first_seen': None,
                'last_seen': None,
                'action': '',
                'from_zone': '',
                'to_zone': '',
                'direction': ''
            }
        
        unique_connections = defaultdict(connection_template)
        
        # Store matching IPs for hostname-based analysis (to determine direction)
        matching_ips = set()
        if self.analysis_mode == "hostname":
            for ip, hostname in self.dns_cache.items():
                if self.target_hostname.lower() in hostname.lower():
                    matching_ips.add(ip)
        
        # Convert DataFrame to dictionary for easier processing
        logs = logs_df.to_dict('records')
        
        # First pass: process connections and build connection records
        print("Processing log entries...")
        for log in logs:
            try:
                # Extract fields with safe defaults
                src_ip = str(log.get('source_ip', ''))
                dst_ip = str(log.get('destination_ip', ''))
                protocol = str(log.get('protocol', ''))
                port = str(log.get('port', ''))
                application = str(log.get('application', ''))
                app_category = str(log.get('app_category', ''))
                app_technology = str(log.get('app_technology', ''))
                app_risk = str(log.get('app_risk', ''))
                rule = str(log.get('rule', ''))
                from_zone = str(log.get('from_zone', ''))
                to_zone = str(log.get('to_zone', ''))
                action = str(log.get('action', ''))
                
                # Get hostnames if already in the logs
                src_hostname = str(log.get('source_hostname', ''))
                dst_hostname = str(log.get('destination_hostname', ''))
                
                # Safe conversion of numeric values
                bytes_total = int(float(log.get('bytes', 0))) if pd.notna(log.get('bytes')) else 0
                bytes_sent = int(float(log.get('bytes_sent', 0))) if pd.notna(log.get('bytes_sent')) else 0
                bytes_received = int(float(log.get('bytes_received', 0))) if pd.notna(log.get('bytes_received')) else 0
                packets = int(float(log.get('packets', 0))) if pd.notna(log.get('packets')) else 0
                
                receive_time = str(log.get('receive_time', ''))
                
                # Create connection key based on analysis mode
                if self.analysis_mode == "ip" and self.target_ip:
                    # IP-based direction
                    if src_ip == self.target_ip:
                        conn_key = f"outbound:{src_ip}:{dst_ip}:{protocol}:{port}:{application}"
                        direction = "outbound"
                    else:
                        conn_key = f"inbound:{src_ip}:{dst_ip}:{protocol}:{port}:{application}"
                        direction = "inbound"
                elif self.analysis_mode == "hostname" and matching_ips:
                    # Hostname-based direction
                    if src_ip in matching_ips:
                        conn_key = f"outbound:{src_ip}:{dst_ip}:{protocol}:{port}:{application}"
                        direction = "outbound"
                    else:
                        conn_key = f"inbound:{src_ip}:{dst_ip}:{protocol}:{port}:{application}"
                        direction = "inbound"
                else:
                    # Zone or all traffic
                    conn_key = f"{from_zone}:{to_zone}:{src_ip}:{dst_ip}:{protocol}:{port}:{application}"
                    direction = f"{from_zone}->{to_zone}"
                
                # Update connection record
                conn = unique_connections[conn_key]
                conn['source_ip'] = src_ip
                conn['destination_ip'] = dst_ip
                
                # Use existing hostname if valid
                if src_hostname and src_hostname not in ('', 'N/A', 'NotFound', 'nan', 'None'):
                    conn['source_hostname'] = src_hostname
                
                if dst_hostname and dst_hostname not in ('', 'N/A', 'NotFound', 'nan', 'None'):
                    conn['destination_hostname'] = dst_hostname
                
                conn['protocol'] = protocol
                conn['port'] = port
                conn['application'] = application
                conn['app_category'] = app_category
                conn['app_technology'] = app_technology
                conn['app_risk'] = app_risk
                conn['rule'] = rule
                conn['hits'] += 1
                conn['bytes'] += bytes_total
                conn['bytes_sent'] += bytes_sent
                conn['bytes_received'] += bytes_received
                conn['packets'] += packets
                conn['direction'] = direction
                conn['action'] = action
                conn['from_zone'] = from_zone
                conn['to_zone'] = to_zone
                
                # Update first/last seen times
                if receive_time:
                    try:
                        # Try both common date formats
                        for fmt in ["%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"]:
                            try:
                                time_obj = datetime.datetime.strptime(receive_time, fmt)
                                if conn['first_seen'] is None or time_obj < conn['first_seen']:
                                    conn['first_seen'] = time_obj
                                if conn['last_seen'] is None or time_obj > conn['last_seen']:
                                    conn['last_seen'] = time_obj
                                break
                            except ValueError:
                                continue
                    except Exception:
                        pass  # Skip on time parsing errors
                    
            except Exception as e:
                print(f"Error processing log entry: {e}")
                continue
        
        # Second pass: Collect all unique IPs for DNS resolution that need it
        source_ips = set()
        dest_ips = set()
        
        print("Collecting unique IPs for DNS resolution...")
        for _, conn in unique_connections.items():
            src_ip = conn['source_ip']
            dst_ip = conn['destination_ip']
            
            # If we haven't already resolved these IPs, add them
            if src_ip and src_ip != 'any' and src_ip not in self.dns_cache:
                source_ips.add(src_ip)
            if dst_ip and dst_ip != 'any' and dst_ip not in self.dns_cache:
                dest_ips.add(dst_ip)
        
        # Resolve DNS in parallel for all collected IPs
        print(f"Collected {len(source_ips)} unique source IPs and {len(dest_ips)} unique destination IPs for resolution")
        
        # Resolve source IPs
        if source_ips:
            source_results = self.resolve_ips_parallel(source_ips)
            self.dns_cache.update(source_results)
        
        # Resolve destination IPs
        if dest_ips:
            dest_results = self.resolve_ips_parallel(dest_ips)
            self.dns_cache.update(dest_results)
        
        # Format and prepare final results
        result = []
        print("Preparing final connection data...")
        
        for _, conn in unique_connections.items():
            # Format datetime objects
            if conn['first_seen']:
                conn['first_seen'] = conn['first_seen'].strftime("%Y/%m/%d %H:%M:%S")
            if conn['last_seen']:
                conn['last_seen'] = conn['last_seen'].strftime("%Y/%m/%d %H:%M:%S")
            
            # Apply resolved hostnames from cache (explicitly ensure we have strings)
            src_ip = conn['source_ip']
            if src_ip in self.dns_cache:
                conn['source_hostname'] = self.dns_cache[src_ip]
            else:
                # Force to "NotFound" string to avoid NaN
                conn['source_hostname'] = "NotFound"
                
            dst_ip = conn['destination_ip']
            if dst_ip in self.dns_cache:
                conn['destination_hostname'] = self.dns_cache[dst_ip]
            else:
                # Force to "NotFound" string to avoid NaN
                conn['destination_hostname'] = "NotFound"
            
            # Double-check we have strings for hostnames, not None or other types
            conn['source_hostname'] = str(conn['source_hostname'])
            conn['destination_hostname'] = str(conn['destination_hostname'])
                
            result.append(conn)
        
        # Print a sample of hostnames for debugging
        if result:
            print("\nSample of resolved hostnames (first 5 connections):")
            for i, conn in enumerate(result[:5]):
                print(f"  {conn['source_ip']} → {conn['source_hostname']}")
                print(f"  {conn['destination_ip']} → {conn['destination_hostname']}")
                if i >= 4:
                    break
        
        return result

    def create_excel_table(self, df, output_file):
        """Create a formatted Excel table from the DataFrame."""
        # Create Excel writer
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # Write DataFrame to Excel
        df.to_excel(writer, sheet_name="Connections", index=False)
        
        # Access the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["Connections"]
        
        # Define table style
        table_style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        
        # Get dimensions and create table
        num_rows = len(df) + 1  # +1 for header
        num_cols = len(df.columns)
        table_range = f"A1:{get_column_letter(num_cols)}{num_rows}"
        table = Table(displayName="ConnectionsTable", ref=table_range)
        table.tableStyleInfo = table_style
        worksheet.add_table(table)
        
        # Set column widths
        for idx, column in enumerate(df.columns, 1):
            col_width = min(max(len(str(column)), df[column].astype(str).str.len().max(), 5), 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = col_width + 2
        
        # Apply header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Save the workbook
        writer.close()

    def generate_report(self, connections, output_file):
        """Generate a CSV or Excel report of unique connections."""
        if not connections:
            print("No connections found to report.")
            return
        
        # Create DataFrame
        df = pd.DataFrame(connections)
        
        print(f"DataFrame created with {len(df)} rows")
        
        # Checking for NaN values in hostname columns
        src_nan_count = df['source_hostname'].isna().sum()
        dst_nan_count = df['destination_hostname'].isna().sum()
        if src_nan_count > 0 or dst_nan_count > 0:
            print(f"WARNING: Found {src_nan_count} NaN values in source_hostname and {dst_nan_count} in destination_hostname")
            print("Fixing NaN values...")
        
        # Extra strong NaN handling for hostname columns
        for col in ['source_hostname', 'destination_hostname']:
            if col in df.columns:
                # First pass: replace NaN with "NotFound"
                df[col] = df[col].fillna('NotFound')
                
                # Second pass: convert all values to strings
                df[col] = df[col].astype(str)
                
                # Third pass: replace any remaining "nan" or empty strings
                df[col] = df[col].replace(['nan', 'None', ''], 'NotFound')
        
        # Handle other columns
        for col in df.columns:
            if col not in ['source_hostname', 'destination_hostname']:
                df[col] = df[col].fillna('')
        
        # Optimize column order
        columns = [
            'direction', 'from_zone', 'to_zone', 
            'source_ip', 'source_hostname', 'destination_ip', 'destination_hostname',
            'protocol', 'port', 'application', 'app_category', 'app_technology', 'app_risk',
            'rule', 'action', 'hits', 'bytes', 'bytes_sent', 'bytes_received', 'packets',
            'first_seen', 'last_seen'
        ]
        
        # Use only columns that exist in the data
        available_columns = [col for col in columns if col in df.columns]
        df = df[available_columns]
        
        # Sort by traffic direction and hits
        sort_columns = ['hits']
        if 'direction' in df.columns:
            sort_columns.insert(0, 'direction')
        df = df.sort_values(sort_columns, ascending=[True, False])
        
        # Save report
        if output_file.endswith('.csv'):
            df.to_csv(output_file, index=False)
        else:
            # Default to Excel if not CSV
            if not output_file.endswith('.xlsx'):
                output_file = output_file + '.xlsx'
            self.create_excel_table(df, output_file)
            
        print(f"Report generated successfully: {output_file}")
        print(f"Found {len(connections)} unique connections.")
        
        # Print summary statistics
        if 'application' in df.columns:
            print("\nTop 10 Applications:")
            top_apps = df.groupby('application')['hits'].sum().sort_values(ascending=False).head(10)
            for app, hits in top_apps.items():
                print(f"  {app}: {hits} hits")
        
        if 'from_zone' in df.columns and 'to_zone' in df.columns:
            print("\nZone Traffic Summary:")
            zone_traffic = df.groupby(['from_zone', 'to_zone'])['hits'].sum().sort_values(ascending=False).head(10)
            for zones, hits in zone_traffic.items():
                print(f"  {zones[0]} -> {zones[1]}: {hits} hits")
        
        if 'action' in df.columns:
            print("\nAction Summary:")
            action_counts = df.groupby('action')['hits'].sum().sort_values(ascending=False)
            for action, hits in action_counts.items():
                print(f"  {action}: {hits} hits")

def main():
    parser = argparse.ArgumentParser(description="Analyze PAN-OS traffic logs from CSV file")
    parser.add_argument("--ip", help="IP endpoint to analyze (optional)")
    parser.add_argument("--zone", help="Zone to analyze (optional)")
    parser.add_argument("--hostname", help="Hostname to analyze (optional)")
    parser.add_argument("--logs", required=True, help="Path to the CSV file with traffic logs")
    parser.add_argument("--output", default="connection_report.xlsx", help="Output file path")
    parser.add_argument("--no-dns", action="store_true", help="Skip DNS resolution")
    parser.add_argument("--dns-timeout", type=float, default=2.0, help="DNS resolution timeout in seconds (default: 2.0)")
    parser.add_argument("--max-workers", type=int, default=50, help="Maximum parallel DNS workers (default: 50)")
    
    args = parser.parse_args()
    
    # Validate arguments - only one filter can be used
    filter_args = [args.ip, args.zone, args.hostname]
    active_filters = [f for f in filter_args if f]
    
    if len(active_filters) > 1:
        print("ERROR: Please specify only one of --ip, --zone, or --hostname")
        sys.exit(1)
    
    if not active_filters:
        print("INFO: No specific filter provided. Analyzing all traffic patterns.")
    
    # Initialize analyzer with target settings
    analyzer = PANOSConnectionAnalyzer(
        target_ip=args.ip, 
        target_zone=args.zone,
        target_hostname=args.hostname,
        dns_timeout=args.dns_timeout,
        max_workers=args.max_workers
    )
    
    # Handle DNS resolution options
    if args.no_dns:
        if args.hostname and analyzer.analysis_mode == "hostname":
            print("WARNING: Cannot disable DNS resolution when using hostname filtering")
        else:
            analyzer.resolve_dns_with_timeout = lambda ip: "DNS-disabled"
            print("DNS resolution disabled")
    
    # Run analysis
    logs_df = analyzer.analyze_logs_from_csv(args.logs)
    connections = analyzer.analyze_connections(logs_df)
    analyzer.generate_report(connections, args.output)

if __name__ == "__main__":
    main()